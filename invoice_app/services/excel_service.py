from pathlib import Path

from openpyxl import load_workbook
import tempfile


class FBRInvoiceLabelNotFound(Exception):
    """Raised when the 'FBR Invoice:' label cannot be found."""
    pass


class FBRInvoiceValueCellNotFound(Exception):
    """Raised when no value cell exists to the right of the label."""
    pass


class ExcelService:

    def update_invoice(self, excel_path, fbr_di):

        workbook = self._load_workbook(excel_path)

        sheet = workbook.active

        label_cell = self._find_label(sheet, "FBR Invoice")

        value_cell = self._find_value_cell(sheet, label_cell)

        self._write_fbr_di(value_cell, fbr_di)

        updated_file = self._save_copy(workbook)

        return updated_file

    def _load_workbook(self, excel_path):

        return load_workbook(excel_path)

    def _find_label(self, sheet, label):

        target = (
            label.lower()
            .replace(":", "")
            .strip()
        )

        for row in sheet.iter_rows():

            for cell in row:

                if cell.value is None:
                    continue

                value = (
                    str(cell.value)
                    .lower()
                    .replace(":", "")
                    .strip()
                )

                if value == target:
                    return cell

        raise FBRInvoiceLabelNotFound(
            f"'{label}' not found."
        )

    def _find_value_cell(self, sheet, label_cell):

        row = label_cell.row

        for col in range(label_cell.column + 1,
                         sheet.max_column + 1):

            cell = sheet.cell(row=row, column=col)

            if cell.value not in (None, ""):
                return cell

        raise FBRInvoiceValueCellNotFound(
            "Could not locate FBR DI cell."
        )

    def _write_fbr_di(self, cell, fbr_di):

        cell.value = fbr_di

    def _save_copy(self, workbook):

        temp_dir = tempfile.gettempdir()

        output_path = Path(temp_dir) / "Updated_Invoice.xlsx"

        workbook.save(output_path)

        return str(output_path)
