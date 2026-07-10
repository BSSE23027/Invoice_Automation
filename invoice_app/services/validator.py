from openpyxl import load_workbook


class InvoiceValidationError(Exception):
    """Raised when the uploaded invoice does not match the latest Gmail invoice."""
    pass


class Validator:

    def validate_invoice_number(self, excel_path, expected_invoice):

        workbook = load_workbook(excel_path, data_only=True)

        sheet = workbook.active

        invoice_number = self._find_invoice_number(sheet)

        if invoice_number is None:
            raise InvoiceValidationError(
                "Invoice number could not be found in the uploaded Excel file."
            )

        if str(invoice_number).strip() != str(expected_invoice).strip():
            raise InvoiceValidationError(
                f"Invoice mismatch. Gmail invoice is {expected_invoice}, "
                f"but uploaded Excel is {invoice_number}."
            )

        return True

    def _find_invoice_number(self, sheet):

        for row in sheet.iter_rows():

            for cell in row:

                if cell.value is None:
                    continue

                value = str(cell.value).strip().lower()

                if value == "dc no":

                    for col in range(
                        cell.column + 1,
                        sheet.max_column + 1
                    ):

                        candidate = sheet.cell(
                            row=cell.row,
                            column=col
                        )

                        if candidate.value not in (None, ""):
                            return candidate.value

        return None
